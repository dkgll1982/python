#!/usr/bin/env python 
# -*- coding:utf-8 -*- 
# @Author: guojun 
# @Company: 航天神舟智慧系统技术有限公司 
# @Site: https://user.qzone.qq.com/350606539/main 
# @Date: 2020-03-07 16:28:44 
# @Last Modified by: guojun 
# @Last Modified time: 2020-03-07 16:28:44 
# @Software: vscode 
# 参考链接：https://blog.csdn.net/magic_engine/article/details/79275341

import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.pyplot import MultipleLocator
import numpy as np
import xlrd
import xlwt
import time

file_nums = ["0112"]
l = ['line_tmp','line_tmpline_tmp2','line_tmpline_tmp3,','line_tmpline_tmp4','line_tmpline_tmp5','line_tmp','line_tmpline_tmp2','line_tmpline_tmp3,','line_tmpline_tmp4','line_tmpline_tmp5','line_tmp','line_tmpline_tmp2','line_tmpline_tmp3,','line_tmpline_tmp4','line_tmpline_tmp5']
start = time.time()
lines = []
try:
    for file_num in file_nums: 
        file_name_ext = 1  
        line_num = 0 
        out_path = r"E:\100-航天智慧\2-源码库\python\backup\excel\人口\\export{}{}.xlsx".format(file_num,"_"+time.strftime('%H%M%S'))
        while line_num<11111:
            print("write all file begin")
            line_num = 0
            for line in range(1): 
                lines.clear()
                lines.append(['a1','b2','c3,','d4','e5','a1','b2','c3,','d4','e5','a1','b2','c3,','d4','e5'])
                for x in range(500000):
                    lines.append(l)
                line_num += 1
                if line_num % 500000 == 0:
                    part_start = time.time()
                    print("write file at :{}".format(part_start))
                    workbook = openpyxl.Workbook(write_only=True)
                    sheet = workbook.create_sheet()

                    for l in lines:
                        sheet.append(l)

                    workbook.save(out_path)
                    workbook.close()
                    workbook = None
                    file_name_ext += 1
                    out_path = r"E:\100-航天智慧\2-源码库\python\backup\excel\人口\\export{}{}.xlsx".format(file_num, "_" +time.strftime('%H%M%S'))

                    part_end = time.time()
                    print("file {} write done at :{}".format(out_path, part_end))
                    print("part used : {}".format(str(part_end - part_start)))
                    lines.clear()

            if lines and len(lines) > 0:
                part_start = time.time()
                print("write file {} at :{}".format(out_path, part_start))
                workbook = openpyxl.Workbook(write_only=True)
                sheet = workbook.create_sheet()

                for l in lines:
                    sheet.append(l)

                out_path = r"E:\100-航天智慧\2-源码库\python\backup\excel\人口\\export{}{}.xlsx".format(file_num,"_"+time.strftime('%H%M%S'))
                workbook.save(out_path)
                workbook.close()
                workbook = None
                part_end = time.time()
                print("file {} write done at :{}".format(out_path, part_end))
                print("part used : {}".format(str(part_end - part_start)))
                lines.clear()

except Exception as e:
    print(e)

end = time.time()
total = end - start
print("write all file finish, used {} times".format(total))